"""
Excel Extractor
===============
Reads .xlsx schedule files using openpyxl (Office Open XML format only).
Legacy binary .xls files (Excel 97-2003) are not supported; they are rejected
at the detector stage with a clear error message.

Sheet selection: uses the active/first sheet by default; only falls back to
the sheet with the most rows when the active sheet is empty.
Converts all cell values to strings matching the CSVExtractor output shape.
Self-registers to ExtractorRegistry on import.
"""
import io
import logging
from pathlib import Path
from typing import Any, Dict, List

import openpyxl

from ingestion.extractors.base import BaseExtractor
from ingestion.extractors.registry import ExtractorRegistry

logger = logging.getLogger(__name__)


def _cell_to_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _pick_sheet(wb: openpyxl.Workbook):
    """
    Return the sheet to extract from.
    Prefers the active (or first) sheet; falls back to the sheet with the
    most rows only when the active sheet has no data rows.
    """
    active = wb.active or (wb.worksheets[0] if wb.worksheets else None)
    if active is None:
        return None

    active_rows = active.max_row or 0
    if active_rows > 1:
        return active

    best = active
    best_rows = active_rows
    for sheet in wb.worksheets:
        if sheet.max_row is not None and sheet.max_row > best_rows:
            best = sheet
            best_rows = sheet.max_row
    return best


def _extract_sheet(ws) -> tuple[List[str], List[List[str]]]:
    """Read headers from first non-empty row; remaining rows become data."""
    all_rows = list(ws.iter_rows(values_only=True))

    headers: List[str] = []
    data_rows: List[List[str]] = []
    header_idx = None

    for i, row in enumerate(all_rows):
        str_row = [_cell_to_str(c) for c in row]
        if any(v for v in str_row):
            headers = str_row
            header_idx = i
            break

    if header_idx is None:
        return [], []

    for row in all_rows[header_idx + 1:]:
        str_row = [_cell_to_str(c) for c in row]
        if any(v for v in str_row):
            data_rows.append(str_row)

    return headers, data_rows


class ExcelExtractor(BaseExtractor):
    def extract(self, file_path: Path) -> Dict[str, Any]:
        return self.extract_from_bytes(file_path.read_bytes(), file_path.name)

    def extract_from_bytes(self, file_bytes: bytes, filename: str) -> Dict[str, Any]:
        try:
            wb = openpyxl.load_workbook(
                filename=io.BytesIO(file_bytes),
                read_only=True,
                data_only=True,
            )
        except Exception as e:
            raise ValueError(f"Cannot open Excel file '{filename}': {e}") from e

        ws = _pick_sheet(wb)
        if ws is None:
            wb.close()
            raise ValueError(f"No worksheets found in '{filename}'.")

        headers, data_rows = _extract_sheet(ws)
        wb.close()

        logger.info(
            f"[{filename}] Excel (.xlsx) extracted: sheet='{ws.title}', "
            f"{len(headers)} columns, {len(data_rows)} data rows"
        )

        raw_text = "\n".join(
            [";".join(headers)] + [";".join(row) for row in data_rows]
        ) if headers else ""

        return {
            "source_system": self.source_system(),
            "headers": headers,
            "rows": data_rows,
            "file_name": filename,
            "raw_text": raw_text,
        }

    def source_system(self) -> str:
        return "EXCEL"


_excel_extractor_instance = ExcelExtractor()
ExtractorRegistry.register("EXCEL", _excel_extractor_instance)
